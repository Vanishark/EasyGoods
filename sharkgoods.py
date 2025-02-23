import ttkbootstrap as tb
from tkinter import filedialog
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.constants import *
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import math

class App(tb.Window):
    def __init__(self):
        super().__init__(
            # themename='superhero'
            )
        self.reshape_window(self)
        self.placewidgets()
        self.title('EasyGoods 谷子排肾表工具 V0.1 Alpha-Test')

    def reshape_window(self,target,scale=1):
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        # 设置窗口宽高为屏幕尺寸的80%
        width = int(screenwidth * 0.7 * scale)
        height = int(screenheight * 0.7 * scale)
        # 计算居中坐标
        x = int((screenwidth - width) / 2)
        y = int((screenheight - height) / 2)
        geometry = f'{width}x{height}+{x}+{y}'
        target.geometry(geometry)

    def placewidgets(self):
         # 顶部按钮区域
        top_frame = tb.Frame(self)
        top_frame.pack(side=tb.TOP, fill=tb.X, padx=10, pady=5)
        
        self.load_btn = tb.Button(top_frame, text="Step 1: 选择xlsx文件", command=self.ask_excel)
        self.load_btn.pack(side=tb.LEFT, padx=5)
        
        self.process_btn = tb.Button(top_frame, text="[x] 进行处理", command=self.process_data,state="disabled")
        self.process_btn.pack(side=tb.LEFT, padx=5)

        self.merge_btn = tb.Button(top_frame, text="[×] 合并入原始表格", command=self.output_merge, state="disabled")
        self.merge_btn.pack(side=tb.LEFT, padx=5)
        
        self.export_btn = tb.Button(top_frame, text="[×] 导出为新表格", command=self.output_new, state="disabled")
        self.export_btn.pack(side=tb.LEFT, padx=5)
        
        # 识别结果报告显示区
        self.report_label = tb.Label(self, text="[↑] 请先点击上方按钮 (Step 1.) 导入Excel。导入成功后会自动识别", justify=tb.LEFT, anchor='w', font=("Consolas", 10))
        self.report_label.pack(fill=tb.X, padx=10, pady=5)

        # 分为两个区域：原始数据预览，汇总输出结果
        mid_frame = tb.Frame(self)
        mid_frame.pack(fill=tb.BOTH, expand=True, padx=10, pady=5)
        mid_frame.columnconfigure(0, weight=1)
        mid_frame.rowconfigure(0, weight=1) 

        # 创建 Notebook 容器，作为 mid_frame 的子控件
        self.notebook = tb.Notebook(mid_frame)
        self.notebook.grid(row=0, column=0, sticky='nsew')  # Notebook 填满整个 mid_frame
        
        importframe = tb.Frame(self.notebook)
        self.notebook.add(importframe, text="导入数据预览")
        exportframe = tb.Frame(self.notebook)
        self.notebook.add(exportframe, text="输出结果预览")

        # 外层 Frame，用于容纳所有设置项
        config_frame = tb.Frame(importframe)
        config_frame.pack(side=TOP, fill=X, padx=10, pady=5)

        self.reportentries={
            "跳过行数":("跳过开头空行，若无为0",tb.StringVar(),'black'),
            "跳过列数":("跳过开头空列，若无为0",tb.StringVar(),'black'),
            "表格标题行":("表格的标题。若无为-1，以蓝色标出。",tb.StringVar(),'blue'),
            "角色行":("包含角色制品的行位置，以绿色标出。",tb.StringVar(),'green'),
            "单价行":("包含角色制品的单价位置，以橙色标出。",tb.StringVar(),'orange'),
            "排表起始行":("包含cn的首行位置，以紫色标出。",tb.StringVar(),'purple'),
        }
        # 创建 5 个设置项，每个设置项都在一个子 Frame 中
        for name, item in self.reportentries.items():
            des, tbsv, color = item
            # 设置项的 Frame
            setting_frame = tb.Frame(config_frame, borderwidth=1, relief="groove", padding=5)
            setting_frame.pack(side=LEFT, padx=5)
            
            # 标签（上方）
            tb.Label(setting_frame, text=name, font=("Consolas", 10, "bold"), foreground=color).pack(side=TOP)
            tb.Label(setting_frame, text=des).pack(side=TOP)
            
            # 输入框（中间）
            tb.Entry(setting_frame, width=5, textvariable=tbsv).pack(side=TOP, pady=2)
            
            # 按钮容器 Frame，用于将加号和减号放在同一行
            button_frame = tb.Frame(setting_frame)
            button_frame.pack(side=TOP, pady=2)
            
            # 加号按钮
            btnup = tb.Button(button_frame, text="+", command=self.identify_and_draw)
            btnup.pack(side=LEFT, padx=2)
            btnup.bind("<Button-1>", lambda e, tv=tbsv: tv.set(str(int(tv.get()) + 1)) if tv.get() else None)
            
            # 减号按钮
            btndown = tb.Button(button_frame, text="-", command=self.identify_and_draw)
            btndown.pack(side=LEFT, padx=2)
            btndown.bind("<Button-1>", lambda e, tv=tbsv: tv.set(str(int(tv.get()) - 1)) if tv.get() else None)

        cols_frame = tb.Frame(exportframe, borderwidth=1, relief="groove", padding=5)
        cols_frame.pack(side=TOP, padx=5)
        
        # 标签（上方）
        tb.Label(cols_frame, text="分栏数", font=("Consolas", 10, "bold"), foreground=color).pack(side=TOP)
        tb.Label(cols_frame, text="将输出表自动分为若干栏，默认为一栏。").pack(side=TOP)
        self.splitcols=tb.StringVar()
        # 输入框（中间）
        tb.Entry(cols_frame, width=5, textvariable=self.splitcols).pack(side=TOP, pady=2)
        self.splitcols.set("1")
        # 按钮容器 Frame，用于将加号和减号放在同一行
        button_frame = tb.Frame(cols_frame)
        button_frame.pack(side=TOP, pady=2)
        
        # 加号按钮
        btnup = tb.Button(button_frame, text="+", command=self.process_data)
        btnup.pack(side=LEFT, padx=2)
        btnup.bind("<Button-1>", lambda e, tv=self.splitcols: tv.set(str(int(tv.get()) + 1)) if tv.get() else None)
        
        # 减号按钮
        btndown = tb.Button(button_frame, text="-", command=self.process_data)
        btndown.pack(side=LEFT, padx=2)
        btndown.bind("<Button-1>", lambda e, tv=self.splitcols: tv.set(str(max(0,int(tv.get()) - 1))) if tv.get() else None)
    

        # 左侧区域（导入数据预览）放在第一个分页里
        self.input_tree = tb.Treeview(importframe)
        self.input_tree.pack(fill=tb.BOTH, expand=True)

        # 右侧区域（输出结果预览）放在第二个分页里
        self.output_tree = tb.Treeview(exportframe)
        self.output_tree.pack(fill=tb.BOTH, expand=True)

    def ask_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        self.originalpath=file_path
        if not file_path:
            return
        try:
            # 读取整个文件（header=None），后续通过预览确认截断
            df = pd.read_excel(file_path, header=None)
            trimmed = self.show_trim_preview(df)
            if trimmed is None:
                Messagebox.show_error("错误", "未确认截断位置！")
                return
            self.original_df = trimmed
            self.file_path = file_path
            self.report_label.config(text="[↓] 导入成功! 请调整下方设置，让系统能正确识别排表，一旦识别正确了，就可以点击(Step 2.)，然后将下方选项卡切换到导出表格预览，检查是否正确。")
            self.reportentries["跳过行数"][1].set("0")
            self.reportentries["跳过列数"][1].set("0")
            self.reportentries["表格标题行"][1].set("0")
            self.reportentries["角色行"][1].set("1")
            self.reportentries["单价行"][1].set("2")
            self.reportentries["排表起始行"][1].set("4")
            self.process_btn.config(text='Step 2.进行处理',state="normal")
            self.identify_and_draw()
            # self.process_data()
        except Exception as e:
            Messagebox.show_error("错误", str(e))
    
    def identify_and_draw(self): 
        # 清空 Treeview 中所有的行
        self.input_tree.delete(*self.input_tree.get_children())

        skippedrows = int(self.reportentries["跳过行数"][1].get())
        skippedcols = int(self.reportentries["跳过列数"][1].get())
        title = int(self.reportentries["表格标题行"][1].get())
        categories = int(self.reportentries["角色行"][1].get())
        prices = int(self.reportentries["单价行"][1].get())
        data_start_row = int(self.reportentries["排表起始行"][1].get())

        # 从原始 DataFrame 中截取数据（假设 self.original_df 已经是一个 pd.DataFrame）
        df: pd.DataFrame = self.original_df.iloc[skippedrows:, skippedcols:]
        
        # 提取标题信息（取标题行第一列）
        try:
            if title != -1:
                title_info = df.iloc[title, 0]
            else:
                title_info = ""
        except Exception:
            title_info = ""
            
        # 提取类别行和价格行（假设从第二列开始）
        try:
            category_row = df.iloc[categories, 1:]
            price_row = df.iloc[prices, 1:]
        except Exception as e:
            Messagebox.show_error("错误", f"解析种类/价格行出错：{e}")
            return
            
        # 提取数据区（排表起始行之后的所有行）
        try:
            data_start_idx = data_start_row - 1  # 转换为 0-indexed
            df_data = df.iloc[data_start_idx:, 1:]
        except Exception as e:
            Messagebox.show_error("错误", f"解析数据区出错：{e}")
            return
            
        # 配置 Treeview 的列（使用 DataFrame 的所有列），并显示 tree 列作为行标题
        df_columns = list(df.columns)[:-1]
        self.input_tree["columns"] = df_columns
        # 同时显示树形列和列标题
        self.input_tree["show"] = "tree headings"
        # 设置 tree 列（#0）的宽度及标题
        self.input_tree.column("#0", width=80, anchor='center')
        self.input_tree.heading("#0", text="列数")
        for col in df_columns:
            self.input_tree.heading(col, text=str(int(col)+1))

        # 配置各类行的颜色标签
        self.input_tree.tag_configure("title_row", background="#ADD8E6")
        self.input_tree.tag_configure("category_row", background="#90EE90")
        self.input_tree.tag_configure("price_row", background="#FFDAB9")
        self.input_tree.tag_configure("data_row", background="#E6E6FA")

        # 插入标题行（如果存在标题信息），并在 tree 列显示“标题行”
        if title_info:
            row_data = [title_info] + [""] * (len(df_columns) - 1)
            self.input_tree.insert("", "end", text="标题行", values=row_data, tags=("title_row",))

        # 插入类别行，在 tree 列显示“类别行”
        row_data = list(category_row)
        if len(row_data) < len(df_columns):
            row_data += [""] * (len(df_columns) - len(row_data))
        self.input_tree.insert("", "end", text="类别行", values=row_data, tags=("category_row",))

        # 插入价格行，在 tree 列显示“价格行”
        row_data = list(price_row)
        if len(row_data) < len(df_columns):
            row_data += [""] * (len(df_columns) - len(row_data))
        self.input_tree.insert("", "end", text="价格行", values=row_data, tags=("price_row",))

        # 插入数据区的每一行，使用 DataFrame 的索引作为行标题
        peishu=1
        for index, row in df_data.iterrows():
            row_data = list(row)
            if len(row_data) < len(df_columns):
                row_data += [""] * (len(df_columns) - len(row_data))
            self.input_tree.insert("", "end", text=f"第{peishu}配", values=row_data, tags=("data_row",))
            peishu=peishu+1
        
    def process_data(self):
        """
        根据当前配置解析 self.original_df，并计算汇总结果，同时更新识别报告显示。
        默认规则：
          - 第一行为标题（通常只在首列有内容）
          - 第二行（索引1）除首列外为种类
          - 第三行（索引2）除首列外为价格
          - 从第四行开始：每行首单元格为配数，后续各列为各人员的订单数量
        """
        skippedrows = int(self.reportentries["跳过行数"][1].get())
        skippedcols = int(self.reportentries["跳过列数"][1].get())
        title = int(self.reportentries["表格标题行"][1].get())
        categories = int(self.reportentries["角色行"][1].get())
        prices = int(self.reportentries["单价行"][1].get())
        data_start_row = int(self.reportentries["排表起始行"][1].get())
        
        df: pd.DataFrame = self.original_df.iloc[skippedrows:, skippedcols:]

        # 提取标题信息（取标题行第一列）
        try:
            if title != -1:
                self.output_title = df.iloc[title, 0]
            else:
                self.output_title = "肾表"
        except Exception:
            self.output_title = "肾表"

        pricetotal={}
        producttotal={}
        for rpos in range(data_start_row-1, df.shape[0]):
            for cpos in range(1, df.shape[1]):
                pricetotal[df.iloc[rpos,cpos]]=pricetotal.get(df.iloc[rpos,cpos],0)+df.iloc[prices,cpos]
                producttotal[df.iloc[rpos,cpos]]=producttotal.get(df.iloc[rpos,cpos],"")+df.iloc[categories,cpos]
        def char_count(s: str) -> str:
            # 使用字典统计各字符的出现次数
            count = {}
            for ch in s:
                count[ch] = count.get(ch, 0) + 1
            # 根据字符首次出现顺序生成结果字符串
            result = []
            seen = set()
            for ch in s:
                if ch not in seen:
                    seen.add(ch)
                    result.append(f"{ch}{count[ch]}")
            return "".join(result)
        for key in producttotal:
            producttotal[key] = char_count(producttotal[key])
        final=pd.DataFrame({'角色制品':producttotal,'应肾':pricetotal})
        final.index.name='cn'
        final.reset_index(drop=False,inplace=True)
        self.final=final
        self.split_cols()

        self.report_label.config(text="[√] 肾表已生成！接下来可以进行分栏输出")
        self.notebook.select(1)
        self.export_btn.config(text='Step 3. 导出为新表格',state="normal")
        self.merge_btn.config(text='Step 3. 合并入原始表格',state="normal")
        #输出到 Output Tree中
        self.output_tree.delete(*self.output_tree.get_children())
        self.output_tree["columns"] = list(self.finaldf.columns)
        self.output_tree["show"] = "headings"  # 仅显示标题，不显示首列

        for i,col in enumerate(self.finaldf.columns.tolist()):
            self.output_tree.heading(i, text=col)

        # 将 DataFrame 中的每一行数据插入到 Treeview 中
        for _, row in self.finaldf.iterrows():
            self.output_tree.insert("", "end", values=list(row))
    
    def show_trim_preview(self, df):
        """
        根据导入的 df 自动识别截断边界：
          - 横向边界：以第二行（索引1）为依据，保留到最后一个非空单元格所在的列（采用1-indexed显示）
          - 竖向边界：对保留的列，取各列最后一个有效值所在行的最小值+1（木桶原则）
        弹出预览窗口，显示截断后的数据，并允许用户修改这两个边界。
        """
        # 横向边界：检查第二行（索引1）
        row2 = df.iloc[1]
        valid_cols = [i for i, x in enumerate(row2) if pd.notna(x)]
        if valid_cols:
            default_h = max(valid_cols) + 1  # 1-indexed数值
        else:
            default_h = df.shape[1]
        
        # 竖向边界：各列最后有效行索引 + 1
        bounds = []
        for i in range(default_h):
            last = df.iloc[:, i].last_valid_index()
            if last is None:
                bounds.append(0)
            else:
                bounds.append(last + 1)
        default_v = min(bounds)
        # 弹出预览窗口（固定尺寸，带滚动条）
        preview_win = tb.Toplevel(self)
        preview_win.title("截断预览与确认")
        self.reshape_window(preview_win,0.6)
        preview_win.grab_set()  # 模态窗口
        tb.Label(preview_win, text="【注意】本步只需要裁切排表右侧边界/下侧边界区域，如果列表开头有空行请在下一步再操作！！").pack(pady=5)
        tb.Label(preview_win, text="系统自动识别的截断位置如下：").pack(pady=5)
        info_text = f"横向截断（列数）：{default_h}    竖向截断（行数）：{default_v}"
        tb.Label(preview_win, text=info_text).pack(pady=5)
        
        # 输入修改区域
        frm = tb.Frame(preview_win)
        frm.pack(pady=5)
        tb.Label(frm, text="→ 横向截断（列数）：").grid(row=0, column=0, padx=5, pady=2)
        h_txtvar=tb.StringVar()
        h_entry = tb.Entry(frm, width=5, textvariable=h_txtvar)
        h_txtvar.set(str(default_h))
        h_entry.grid(row=0, column=1, padx=5, pady=2)
        
        h_upbtn=tb.Button(frm, text='+', command=lambda: update_preview())
        h_upbtn.grid(row=0, column=2, padx=5, pady=2)
        h_downbtn=tb.Button(frm, text='-', command=lambda: update_preview())
        h_downbtn.grid(row=0, column=3, padx=5, pady=2)
        
        tb.Label(frm, text="↓ 竖向截断（行数）：").grid(row=1, column=0, padx=5, pady=2)
        v_txtvar=tb.StringVar()
        v_entry = tb.Entry(frm, width=5, textvariable=v_txtvar)
        v_txtvar.set(str(default_v))
        v_entry.grid(row=1, column=1, padx=5, pady=2)

        v_upbtn = tb.Button(frm, text='+', command=lambda: update_preview())
        v_upbtn.grid(row=1, column=2, padx=5, pady=2)
        v_downbtn = tb.Button(frm, text='-', command=lambda: update_preview())
        v_downbtn.grid(row=1, column=3, padx=5, pady=2)
        
        # 预览区
        preview_frame = tb.Frame(preview_win)
        preview_frame.pack(fill=tb.BOTH, expand=True, padx=10, pady=5)
        tree = tb.Treeview(preview_frame, height=10)
        tree.pack(side=tb.LEFT, fill=tb.BOTH, expand=True)
        
        def update_preview():
            try:
                new_h = int(h_txtvar.get())
                new_v = int(v_txtvar.get())
            except Exception:
                return
            # 构造一个尺寸为 new_v x new_h 的预览 DataFrame，初始全部为 nan
            import numpy as np
            preview_df = pd.DataFrame(np.nan, index=range(new_v), columns=range(new_h))

            # 确定原始数据与用户选区的交集区域
            rows = min(new_v, df.shape[0])
            cols = min(new_h, df.shape[1])
            
            # 将原始数据中对应区域拷贝到预览 DataFrame
            preview_df.iloc[:rows, :cols] = df.iloc[:rows, :cols]
            tree.delete(*tree.get_children())
            cols_list = [str(i+1) for i in range(new_h)]
            tree["columns"] = cols_list
            # 同时显示树形列和表头，这样就能看到行标题了
            tree["show"] = "tree headings"

            # 设置树形列（#0）作为行标题
            tree.heading("#0", text="Row")       # 行标题名称
            tree.column("#0", width=80, anchor='center')

            # 设置各列标题
            for col in cols_list:
                tree.heading(col, text=col)
                tree.column(col, width=80, anchor='center')

            # 插入数据时，传入行标题到 text 参数
            for i, row in preview_df.iterrows():
                row_values = [("" if pd.isna(x) else x) for x in row]
                tree.insert("", "end", text=str(i+1), values=row_values)

        update_preview()
        h_entry.bind("<KeyRelease>", lambda e: update_preview())
        v_entry.bind("<KeyRelease>", lambda e: update_preview())
        h_downbtn.bind("<ButtonRelease-1>", lambda e: h_txtvar.set(str(int(h_txtvar.get())-1)))
        h_upbtn.bind("<Button-1>", lambda e: h_txtvar.set(str(int(h_txtvar.get())+1)))
        v_downbtn.bind("<ButtonRelease-1>", lambda e: v_txtvar.set(str(int(v_txtvar.get())-1)))
        v_upbtn.bind("<ButtonRelease-1>", lambda e: v_txtvar.set(str(int(v_txtvar.get())+1)))
        
        result = {}
        def confirm():
            try:
                new_h = int(h_txtvar.get())
                new_v = int(v_txtvar.get())
            except Exception:
                Messagebox.show_error("错误", "请输入有效数字！")
                return
            result['h_bound'] = new_h
            result['v_bound'] = new_v
            preview_win.destroy()
        
        btn_frame = tb.Frame(preview_win)
        btn_frame.pack(pady=5)
        confirm_btn = tb.Button(btn_frame, text="确认截断位置", command=confirm)
        confirm_btn.pack()
        self.wait_window(preview_win)
        if 'h_bound' in result.keys() and 'v_bound' in result.keys():
            trimmed_df = df.iloc[:result['v_bound'], :result['h_bound']]
            return trimmed_df
        else:
            return None

    def getrepeats(self):
        try:
            scols=max(1,int(self.splitcols.get()))
        except Exception as e:
            # print(e)
            scols=1
        return scols
    def split_cols(self):
        scols=self.getrepeats()
        # 获取 DataFrame 的总行数
        total_rows = self.final.shape[0]
        # 每份的行数，向上取整
        rows_per_part = math.ceil(total_rows / scols)
        parts = []
        for i in range(scols):
            start = i * rows_per_part
            end = start + rows_per_part
            # 分块，并重置索引，确保横向拼接时对齐
            if i==scols-1:
                part = self.final.iloc[start:].reset_index(drop=True)
            else:
                part = self.final.iloc[start:end].reset_index(drop=True)
            parts.append(part)
        
        # 横向拼接所有分块，axis=1 表示按列拼接
        self.finaldf = pd.concat(parts, axis=1)

    def output_merge(self):
        # 弹出系统文件保存对话框，选择导出文件的路径和名称（扩展名为 .xlsx）
        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="导出合并文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        try:
            # 从原始文件路径加载工作簿
            wb = load_workbook(self.originalpath)
            ws = wb.active

            dim = ws.calculate_dimension()
            if ":" in dim:
                end_cell = dim.split(":")[1]
                # 过滤出列字母部分，再转换为列索引
                orig_max_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
            else:
                orig_max_col = 1

            # 新增部分在原始表格右侧加入，空2列后开始写入
            new_start_col = orig_max_col + 2
            # 定义细边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 取得 self.finaldf 的列数
            final_num_cols = len(self.finaldf.columns)
            merged_end_col = new_start_col + final_num_cols - 1

            # 在第一行写入标题，并合并新增区域（新部分标题区域）
            title_text = self.output_title  # 假定 self.output_title 已赋值
            ws.merge_cells(start_row=1, start_column=new_start_col, end_row=1, end_column=merged_end_col)
            title_cell = ws.cell(row=1, column=new_start_col)
            title_cell.value = title_text
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = PatternFill(fill_type='solid', fgColor="FFA500")
            # 为合并区域内每个单元格添加边框
            for col in range(new_start_col, merged_end_col + 1):
                ws.cell(row=1, column=col).border = thin_border

            # 写入新增部分的表头（写在第二行）
            for idx, col_name in enumerate(self.finaldf.columns, start=new_start_col):
                cell = ws.cell(row=2, column=idx)
                cell.value = col_name
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
                cell.border = thin_border

            # 写入数据行（从第三行开始）
            for row_idx, row in enumerate(self.finaldf.itertuples(index=False), start=3):
                for col_offset, value in enumerate(row, start=new_start_col):
                    cell = ws.cell(row=row_idx, column=col_offset)
                    cell.value = value
                    cell.border = thin_border

                    # 根据列名判断是否需要设置背景颜色
                    col_name = self.finaldf.columns[col_offset - new_start_col]
                    if col_name.lower() == "cn":
                        cell.fill = PatternFill(fill_type='solid', fgColor="90EE90")
                    elif col_name == "应肾":
                        cell.fill = PatternFill(fill_type='solid', fgColor="FFFF00")

            # 可选：调整新增部分各列宽度
            for col in range(new_start_col, merged_end_col + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 15

            wb.save(file_path)
            Messagebox.show_info(f"文件成功导出至：\n{file_path}", "成功", parent=self)
        except Exception as e:
            Messagebox.show_error(f"保存文件时出错：{e}", "导出失败", parent=self)
    

    def output_new(self):
        # 弹出保存文件对话框，选择导出文件的路径和名称（扩展名为 .xlsx）
        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="导出文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active

            # DataFrame 的列数，用于后续写入和合并操作
            num_cols = len(self.finaldf.columns)

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # 在第一行写入 self.output_title，并合并所有列（A1 到最后一列）
            title_text = self.output_title  # 假定 self.output_title 已赋值
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = title_text
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.font = Font(bold=True, size=14)
            # 设置标题行背景色为橙色（十六进制颜色代码：FFA500）
            title_cell.fill = PatternFill(fill_type='solid', fgColor="FFA500")
            for col in range(1, num_cols + 1):
                ws.cell(row=1, column=col).border = thin_border
            # 写入表头（第二行），遍历 DataFrame 的列名
            for col_idx, col_name in enumerate(self.finaldf.columns, start=1):
                cell = ws.cell(row=2, column=col_idx)
                cell.value = col_name
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
                cell.border = thin_border

            # 写入数据行（从第三行开始）
            for row_idx, row in enumerate(self.finaldf.itertuples(index=False), start=3):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = thin_border

                    # 根据列名判断是否需要设置背景颜色
                    col_name = self.finaldf.columns[col_idx - 1]
                    if col_name == "cn":
                        # 设置 "CN" 列的数据背景为浅绿色（例如：90EE90）
                        cell.fill = PatternFill(fill_type='solid', fgColor="90EE90")
                    elif col_name == "应肾":
                        # 设置 "应肾" 列的数据背景为黄色（例如：FFFF00）
                        cell.fill = PatternFill(fill_type='solid', fgColor="FFFF00")

            # 可选：调整各列宽度，方便查看
            for col_idx in range(1, num_cols + 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 15

            # 保存 Excel 文件
            wb.save(file_path)
            Messagebox.show_info(f"文件成功导出至：\n{file_path}","成功", parent=self)
        except Exception as e:
            Messagebox.show_error(f"保存文件时出错：{e}","导出失败", parent=self)
if __name__=='__main__':
    import warnings
    warnings.filterwarnings("ignore")
    app = App()
    try:
        app.mainloop()
    except Exception as e:
        Messagebox.show_error(f"程序出错：{e}","错误", parent=app)