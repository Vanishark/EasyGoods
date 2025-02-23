import ttkbootstrap as tb
from tkinter import filedialog
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.constants import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import math

class App(tb.Window):
    def __init__(self):
        super().__init__()
        self.reshape_window(self)
        self.placewidgets()
        self.title('EasyGoods 谷子排肾表工具 V1.0 Alpha-Test')

    def reshape_window(self, target, scale=1):
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        width = int(screenwidth * 0.7 * scale)
        height = int(screenheight * 0.7 * scale)
        x = int((screenwidth - width) / 2)
        y = int((screenheight - height) / 2)
        geometry = f'{width}x{height}+{x}+{y}'
        target.geometry(geometry)

    def placewidgets(self):
        top_frame = tb.Frame(self)
        top_frame.pack(side=tb.TOP, fill=tb.X, padx=10, pady=5)
        
        self.load_btn = tb.Button(top_frame, text="Step 1: 选择xlsx文件", command=self.ask_excel)
        self.load_btn.pack(side=tb.LEFT, padx=5)
        
        self.process_btn = tb.Button(top_frame, text="[x] 进行处理", command=self.process_data, state="disabled")
        self.process_btn.pack(side=tb.LEFT, padx=5)

        self.merge_btn = tb.Button(top_frame, text="[×] 合并入原始表格", command=self.output_merge, state="disabled")
        self.merge_btn.pack(side=tb.LEFT, padx=5)
        
        self.export_btn = tb.Button(top_frame, text="[×] 导出为新表格", command=self.output_new, state="disabled")
        self.export_btn.pack(side=tb.LEFT, padx=5)
        
        self.report_label = tb.Label(self, text="[↑] 请先点击上方按钮 (Step 1.) 导入Excel。导入成功后会自动识别", 
                                     justify=tb.LEFT, anchor='w', font=("Consolas", 10))
        self.report_label.pack(fill=tb.X, padx=10, pady=5)

        mid_frame = tb.Frame(self)
        mid_frame.pack(fill=tb.BOTH, expand=True, padx=10, pady=5)
        mid_frame.columnconfigure(0, weight=1)
        mid_frame.rowconfigure(0, weight=1) 

        self.notebook = tb.Notebook(mid_frame)
        self.notebook.grid(row=0, column=0, sticky='nsew')
        
        importframe = tb.Frame(self.notebook)
        self.notebook.add(importframe, text="导入数据预览")
        exportframe = tb.Frame(self.notebook)
        self.notebook.add(exportframe, text="输出结果预览")

        config_frame = tb.Frame(importframe)
        config_frame.pack(side=tb.TOP, fill=tb.X, padx=10, pady=5)

        self.reportentries = {
            "跳过行数":("跳过开头空行，若无为0", tb.StringVar(), 'black'),
            "跳过列数":("跳过开头空列，若无为0", tb.StringVar(), 'black'),
            "表格标题行":("表格的标题。若无为-1，以蓝色标出。", tb.StringVar(), 'blue'),
            "角色行":("包含角色制品的行位置，以绿色标出。", tb.StringVar(), 'green'),
            "单价行":("包含角色制品的单价位置，以橙色标出。", tb.StringVar(), 'orange'),
            "排表起始行":("包含cn的首行位置，以紫色标出。", tb.StringVar(), 'purple'),
        }
        for name, item in self.reportentries.items():
            des, tbsv, color = item
            setting_frame = tb.Frame(config_frame, borderwidth=1, relief="groove", padding=5)
            setting_frame.pack(side=tb.LEFT, padx=5)
            
            tb.Label(setting_frame, text=name, font=("Consolas", 10, "bold"), foreground=color).pack(side=tb.TOP)
            tb.Label(setting_frame, text=des).pack(side=tb.TOP)
            tb.Entry(setting_frame, width=5, textvariable=tbsv).pack(side=tb.TOP, pady=2)
            
            button_frame = tb.Frame(setting_frame)
            button_frame.pack(side=tb.TOP, pady=2)
            
            btnup = tb.Button(button_frame, text="+", command=self.identify_and_draw)
            btnup.pack(side=tb.LEFT, padx=2)
            btnup.bind("<Button-1>", lambda e, tv=tbsv: tv.set(str(int(tv.get()) + 1)) if tv.get() else None)
            
            btndown = tb.Button(button_frame, text="-", command=self.identify_and_draw)
            btndown.pack(side=tb.LEFT, padx=2)
            btndown.bind("<Button-1>", lambda e, tv=tbsv: tv.set(str(int(tv.get()) - 1)) if tv.get() else None)

        cols_frame = tb.Frame(exportframe, borderwidth=1, relief="groove", padding=5)
        cols_frame.pack(side=tb.TOP, padx=5)
        
        tb.Label(cols_frame, text="分栏数", font=("Consolas", 10, "bold")).pack(side=tb.TOP)
        tb.Label(cols_frame, text="将输出表自动分为若干栏，默认为一栏。").pack(side=tb.TOP)
        self.splitcols = tb.StringVar()
        tb.Entry(cols_frame, width=5, textvariable=self.splitcols).pack(side=tb.TOP, pady=2)
        self.splitcols.set("1")
        button_frame = tb.Frame(cols_frame)
        button_frame.pack(side=tb.TOP, pady=2)
        
        btnup = tb.Button(button_frame, text="+", command=self.process_data)
        btnup.pack(side=tb.LEFT, padx=2)
        btnup.bind("<Button-1>", lambda e, tv=self.splitcols: tv.set(str(int(tv.get()) + 1)) if tv.get() else None)
        
        btndown = tb.Button(button_frame, text="-", command=self.process_data)
        btndown.pack(side=tb.LEFT, padx=2)
        btndown.bind("<Button-1>", lambda e, tv=self.splitcols: tv.set(str(max(0, int(tv.get()) - 1))) if tv.get() else None)
    
        self.input_tree = tb.Treeview(importframe)
        self.input_tree.pack(fill=tb.BOTH, expand=True)

        self.output_tree = tb.Treeview(exportframe)
        self.output_tree.pack(fill=tb.BOTH, expand=True)

    def ask_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        self.originalpath = file_path
        if not file_path:
            return
        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active
            data = list(sheet.values)
            trimmed = self.show_trim_preview(data)
            if trimmed is None:
                Messagebox.show_error("错误", "未确认截断位置！")
                return
            self.original_data = trimmed
            self.file_path = file_path
            self.report_label.config(text="[↓] 导入成功! 请调整下方设置，让系统能正确识别排表，一旦识别正确了，就可以点击(Step 2.)，然后将下方选项卡切换到导出表格预览，检查是否正确。")
            self.reportentries["跳过行数"][1].set("0")
            self.reportentries["跳过列数"][1].set("0")
            self.reportentries["表格标题行"][1].set("0")
            self.reportentries["角色行"][1].set("1")
            self.reportentries["单价行"][1].set("2")
            self.reportentries["排表起始行"][1].set("4")
            self.process_btn.config(text='Step 2.进行处理', state="normal")
            self.identify_and_draw()
        except Exception as e:
            Messagebox.show_error("错误", str(e))
    
    def show_trim_preview(self, data):
        # 横向边界：检查第二行
        row2 = data[1] if len(data) > 1 else []
        valid_cols = [i for i, x in enumerate(row2) if x is not None]
        if valid_cols:
            default_h = max(valid_cols) + 1
        else:
            default_h = len(data[0]) if data else 0
        bounds = []
        for i in range(default_h):
            last = -1
            for j, row in enumerate(data):
                if i < len(row) and row[i] is not None:
                    last = j
            bounds.append(last + 1 if last >= 0 else 0)
        default_v = min(bounds) if bounds else 0
        preview_win = tb.Toplevel(self)
        preview_win.title("截断预览与确认")
        self.reshape_window(preview_win, 0.6)
        preview_win.grab_set()
        tb.Label(preview_win, text="【注意】本步只需要裁切排表右侧边界/下侧边界区域，如果列表开头有空行请在下一步再操作！！").pack(pady=5)
        tb.Label(preview_win, text="系统自动识别的截断位置如下：").pack(pady=5)
        info_text = f"横向截断（列数）：{default_h}    竖向截断（行数）：{default_v}"
        tb.Label(preview_win, text=info_text).pack(pady=5)
        
        frm = tb.Frame(preview_win)
        frm.pack(pady=5)
        tb.Label(frm, text="→ 横向截断（列数）：").grid(row=0, column=0, padx=5, pady=2)
        h_txtvar = tb.StringVar()
        h_entry = tb.Entry(frm, width=5, textvariable=h_txtvar)
        h_txtvar.set(str(default_h))
        h_entry.grid(row=0, column=1, padx=5, pady=2)
        
        h_upbtn = tb.Button(frm, text='+', command=lambda: update_preview())
        h_upbtn.grid(row=0, column=2, padx=5, pady=2)
        h_downbtn = tb.Button(frm, text='-', command=lambda: update_preview())
        h_downbtn.grid(row=0, column=3, padx=5, pady=2)
        
        tb.Label(frm, text="↓ 竖向截断（行数）：").grid(row=1, column=0, padx=5, pady=2)
        v_txtvar = tb.StringVar()
        v_entry = tb.Entry(frm, width=5, textvariable=v_txtvar)
        v_txtvar.set(str(default_v))
        v_entry.grid(row=1, column=1, padx=5, pady=2)
        
        v_upbtn = tb.Button(frm, text='+', command=lambda: update_preview())
        v_upbtn.grid(row=1, column=2, padx=5, pady=2)
        v_downbtn = tb.Button(frm, text='-', command=lambda: update_preview())
        v_downbtn.grid(row=1, column=3, padx=5, pady=2)
        
        preview_frame = tb.Frame(preview_win)
        preview_frame.pack(fill=tb.BOTH, expand=True, padx=10, pady=5)
        tree = tb.Treeview(preview_frame, height=10)
        tree.pack(side=tb.LEFT, fill=tb.BOTH, expand=True)
        
        def update_preview():
            try:
                new_h = int(h_txtvar.get())
                new_v = int(v_txtvar.get())
            except Exception:
                return
            preview_data = [["" for _ in range(new_h)] for _ in range(new_v)]
            rows = min(new_v, len(data))
            cols = min(new_h, len(data[0]) if data else 0)
            for i in range(rows):
                for j in range(cols):
                    preview_data[i][j] = data[i][j]
            tree.delete(*tree.get_children())
            cols_list = [str(i+1) for i in range(new_h)]
            tree["columns"] = cols_list
            tree["show"] = "tree headings"
            tree.heading("#0", text="Row")
            tree.column("#0", width=80, anchor='center')
            for col in cols_list:
                tree.heading(col, text=col)
                tree.column(col, width=80, anchor='center')
            for i, row in enumerate(preview_data):
                row_values = [("" if (x is None) else x) for x in row]
                tree.insert("", "end", text=str(i+1), values=row_values)
        
        update_preview()
        h_entry.bind("<KeyRelease>", lambda e: update_preview())
        v_entry.bind("<KeyRelease>", lambda e: update_preview())
        h_downbtn.bind("<ButtonRelease-1>", lambda e: h_txtvar.set(str(int(h_txtvar.get())-1)))
        h_upbtn.bind("<Button-1>", lambda e: h_txtvar.set(str(int(h_txtvar.get())+1)))
        v_downbtn.bind("<ButtonRelease-1>", lambda e: v_txtvar.set(str(int(v_txtvar.get())-1)))
        v_upbtn.bind("<ButtonRelease-1>", lambda e: v_txtvar.set(str(int(v_txtvar.get())+1)))
        
        result = {}
        def confirm():
            try:
                new_h = int(h_txtvar.get())
                new_v = int(v_txtvar.get())
            except Exception:
                Messagebox.show_error("错误", "请输入有效数字！")
                return
            result['h_bound'] = new_h
            result['v_bound'] = new_v
            preview_win.destroy()
        
        btn_frame = tb.Frame(preview_win)
        btn_frame.pack(pady=5)
        confirm_btn = tb.Button(btn_frame, text="确认截断位置", command=confirm)
        confirm_btn.pack()
        self.wait_window(preview_win)
        if 'h_bound' in result and 'v_bound' in result:
            trimmed_data = [row[:result['h_bound']] for row in data[:result['v_bound']]]
            return trimmed_data
        else:
            return None

    def identify_and_draw(self):
        self.input_tree.delete(*self.input_tree.get_children())
        skippedrows = int(self.reportentries["跳过行数"][1].get())
        skippedcols = int(self.reportentries["跳过列数"][1].get())
        title = int(self.reportentries["表格标题行"][1].get())
        categories = int(self.reportentries["角色行"][1].get())
        prices = int(self.reportentries["单价行"][1].get())
        data_start_row = int(self.reportentries["排表起始行"][1].get())
        
        df = [row[skippedcols:] for row in self.original_data[skippedrows:]]
        
        try:
            if title != -1 and title < len(df) and len(df[title]) > 0:
                title_info = df[title][0]
            else:
                title_info = ""
        except Exception:
            title_info = ""
            
        try:
            category_row = df[categories][1:] if categories < len(df) else []
            price_row = df[prices][1:] if prices < len(df) else []
        except Exception as e:
            Messagebox.show_error("错误", f"解析种类/价格行出错：{e}")
            return
            
        try:
            data_start_idx = data_start_row - 1
            df_data = []
            for row in df[data_start_idx:]:
                df_data.append(row[1:] if len(row) > 1 else [])
        except Exception as e:
            Messagebox.show_error("错误", f"解析数据区出错：{e}")
            return
            
        num_cols = len(df[0]) - 1 if df and len(df[0]) > 1 else 0
        df_columns = list(range(num_cols))
        self.input_tree["columns"] = df_columns
        self.input_tree["show"] = "tree headings"
        self.input_tree.column("#0", width=80, anchor='center')
        self.input_tree.heading("#0", text="列数")
        for col in df_columns:
            self.input_tree.heading(col, text=str(col+1))
        self.input_tree.tag_configure("title_row", background="#ADD8E6")
        self.input_tree.tag_configure("category_row", background="#90EE90")
        self.input_tree.tag_configure("price_row", background="#FFDAB9")
        self.input_tree.tag_configure("data_row", background="#E6E6FA")
        
        if title_info:
            row_data = [title_info] + ["" for _ in range(num_cols - 1)]
            self.input_tree.insert("", "end", text="标题行", values=row_data, tags=("title_row",))
        
        row_data = list(category_row)
        if len(row_data) < num_cols:
            row_data += ["" for _ in range(num_cols - len(row_data))]
        self.input_tree.insert("", "end", text="类别行", values=row_data, tags=("category_row",))
        
        row_data = list(price_row)
        if len(row_data) < num_cols:
            row_data += ["" for _ in range(num_cols - len(row_data))]
        self.input_tree.insert("", "end", text="价格行", values=row_data, tags=("price_row",))
        
        peishu = 1
        for row in df_data:
            row_data = list(row)
            if len(row_data) < num_cols:
                row_data += ["" for _ in range(num_cols - len(row_data))]
            self.input_tree.insert("", "end", text=f"第{peishu}配", values=row_data, tags=("data_row",))
            peishu += 1
        
    def process_data(self):
        skippedrows = int(self.reportentries["跳过行数"][1].get())
        skippedcols = int(self.reportentries["跳过列数"][1].get())
        title = int(self.reportentries["表格标题行"][1].get())
        categories = int(self.reportentries["角色行"][1].get())
        prices = int(self.reportentries["单价行"][1].get())
        data_start_row = int(self.reportentries["排表起始行"][1].get())
        
        df = [row[skippedcols:] for row in self.original_data[skippedrows:]]
        try:
            if title != -1 and title < len(df) and len(df[title]) > 0:
                self.output_title = df[title][0]
            else:
                self.output_title = "肾表"
        except Exception:
            self.output_title = "肾表"

        pricetotal = {}
        producttotal = {}
        for rpos in range(data_start_row - 1, len(df)):
            row = df[rpos]
            for cpos in range(1, len(row)):
                key = row[cpos]
                if key is None:
                    continue
                price_val = df[prices][cpos] if prices < len(df) and cpos < len(df[prices]) and df[prices][cpos] is not None else 0
                pricetotal[key] = pricetotal.get(key, 0) + price_val
                cat_val = df[categories][cpos] if categories < len(df) and cpos < len(df[categories]) and df[categories][cpos] is not None else ""
                producttotal[key] = producttotal.get(key, "") + str(cat_val)
        def char_count(s: str) -> str:
            seen = set()
            out = []
            for ch in s:
                if ch not in seen:
                    seen.add(ch)
                    out.append(f"{ch}{s.count(ch)}")
            return "".join(out)
        for key in producttotal:
            producttotal[key] = char_count(producttotal[key])
        final_table = []
        for key in producttotal:
            final_table.append([key, producttotal[key], pricetotal.get(key, 0)])
        self.final = final_table
        self.final_header = ["cn", "角色制品", "应肾"]
        self.split_cols()
        
        self.report_label.config(text="[√] 肾表已生成！接下来可以进行分栏输出")
        self.notebook.select(1)
        self.export_btn.config(text='Step 3. 导出为新表格', state="normal")
        self.merge_btn.config(text='Step 3. 合并入原始表格', state="normal")
        self.output_tree.delete(*self.output_tree.get_children())
        if hasattr(self, "finaldf") and "columns" in self.finaldf:
            columns = self.finaldf["columns"]
        else:
            columns = self.final_header
        self.output_tree["columns"] = columns
        self.output_tree["show"] = "headings"
        for i, col in enumerate(columns):
            self.output_tree.heading(i, text=str(col))
        for row in self.finaldf["data"]:
            self.output_tree.insert("", "end", values=row)
    
    def split_cols(self):
        try:
            scols = max(1, int(self.splitcols.get()))
        except Exception:
            scols = 1
        total_rows = len(self.final)
        rows_per_part = math.ceil(total_rows / scols)
        parts = []
        for i in range(scols):
            start = i * rows_per_part
            end = start + rows_per_part
            part = self.final[start:end]
            parts.append(part)
        num_cols_per_part = len(self.final_header)
        new_data = []
        max_rows = max(len(part) for part in parts) if parts else 0
        for part in parts:
            while len(part) < max_rows:
                part.append([""] * num_cols_per_part)
        for i in range(max_rows):
            new_row = []
            for part in parts:
                new_row.extend(part[i])
            new_data.append(new_row)
        new_header = []
        for _ in range(scols):
            new_header.extend(self.final_header)
        self.finaldf = {"columns": new_header, "data": new_data}

    def output_merge(self):
        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="导出合并文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        try:
            wb = load_workbook(self.originalpath)
            ws = wb.active

            dim = ws.calculate_dimension()
            if ":" in dim:
                end_cell = dim.split(":")[1]
                orig_max_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
            else:
                orig_max_col = 1

            new_start_col = orig_max_col + 2
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            final_num_cols = len(self.finaldf["columns"])
            merged_end_col = new_start_col + final_num_cols - 1

            title_text = self.output_title
            ws.merge_cells(start_row=1, start_column=new_start_col, end_row=1, end_column=merged_end_col)
            title_cell = ws.cell(row=1, column=new_start_col)
            title_cell.value = title_text
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = PatternFill(fill_type='solid', fgColor="FFA500")
            for col in range(new_start_col, merged_end_col + 1):
                ws.cell(row=1, column=col).border = thin_border

            for idx, col_name in enumerate(self.finaldf["columns"], start=new_start_col):
                cell = ws.cell(row=2, column=idx)
                cell.value = col_name
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
                cell.border = thin_border

            for row_idx, row in enumerate(self.finaldf["data"], start=3):
                for col_offset, value in enumerate(row, start=new_start_col):
                    cell = ws.cell(row=row_idx, column=col_offset)
                    cell.value = value
                    cell.border = thin_border
                    col_name = self.finaldf["columns"][col_offset - new_start_col]
                    if isinstance(col_name, str) and col_name.lower() == "cn":
                        cell.fill = PatternFill(fill_type='solid', fgColor="90EE90")
                    elif col_name == "应肾":
                        cell.fill = PatternFill(fill_type='solid', fgColor="FFFF00")

            for col in range(new_start_col, merged_end_col + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 15

            wb.save(file_path)
            Messagebox.show_info(f"文件成功导出至：\n{file_path}", "成功", parent=self)
        except Exception as e:
            Messagebox.show_error(f"保存文件时出错：{e}", "导出失败", parent=self)
    
    def output_new(self):
        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="导出文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active

            num_cols = len(self.finaldf["columns"])
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            title_text = self.output_title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = title_text
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = PatternFill(fill_type='solid', fgColor="FFA500")
            for col in range(1, num_cols + 1):
                ws.cell(row=1, column=col).border = thin_border
            for col_idx, col_name in enumerate(self.finaldf["columns"], start=1):
                cell = ws.cell(row=2, column=col_idx)
                cell.value = col_name
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
                cell.border = thin_border
            for row_idx, row in enumerate(self.finaldf["data"], start=3):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    col_name = self.finaldf["columns"][col_idx - 1]
                    if isinstance(col_name, str) and col_name.lower() == "cn":
                        cell.fill = PatternFill(fill_type='solid', fgColor="90EE90")
                    elif col_name == "应肾":
                        cell.fill = PatternFill(fill_type='solid', fgColor="FFFF00")
            for col_idx in range(1, num_cols + 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 15

            wb.save(file_path)
            Messagebox.show_info(f"文件成功导出至：\n{file_path}", "成功", parent=self)
        except Exception as e:
            Messagebox.show_error(f"保存文件时出错：{e}", "导出失败", parent=self)

if __name__=='__main__':
    import warnings
    warnings.filterwarnings("ignore")
    app = App()
    try:
        app.mainloop()
    except Exception as e:
        Messagebox.show_error(f"程序出错：{e}", "错误", parent=app)
